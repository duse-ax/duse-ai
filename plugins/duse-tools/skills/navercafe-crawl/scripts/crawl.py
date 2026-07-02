#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
navercafe-crawl : 네이버 카페 게시판 전체 게시글 크롤링 → 엑셀.

사용:
  python3 crawl.py --url "<게시판 URL>" --cookie cookie_header.txt --out "결과.xlsx"

전제:
  - list(목록) API 는 비로그인으로 조회 가능하지만, 본문/댓글은 로그인 쿠키 필요.
  - 쿠키를 제공하는 계정이 해당 카페의 "회원"이어야 본문이 보인다(비회원이면 401/403).
  - cookie_header.txt : "NID_AUT=...; NID_SES=...; NNB=...; nid_inf=..." 한 줄.

의존: 표준 라이브러리 + beautifulsoup4 + lxml + openpyxl
"""
import argparse, json, os, re, sys, time, html as htmlmod
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone, timedelta
import urllib.request, urllib.error

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

KST = timezone(timedelta(hours=9))
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0 Safari/537.36")
APIS = "https://apis.naver.com/cafe-web"


# ---------------------------------------------------------------- URL 파싱
def parse_url(url):
    """https://cafe.naver.com/f-e/cafes/24646884/menus/30  ->  (24646884, 30)
    m.cafe.naver.com/ca-fe/web/cafes/{id}/menus/{menu} 형태도 지원."""
    m = re.search(r"/cafes/(\d+)/menus/(\d+)", url)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.search(r"clubid=(\d+).*?menuid=(\d+)", url, re.I)
    if m:
        return int(m.group(1)), int(m.group(2))
    raise SystemExit(
        "URL에서 카페ID/메뉴ID를 못 찾았습니다. 예: "
        "https://cafe.naver.com/f-e/cafes/24646884/menus/30")


# ---------------------------------------------------------------- HTTP
def http_get(url, cookie=None, referer=None, retries=3, timeout=25):
    headers = {"User-Agent": UA}
    if cookie:
        headers["Cookie"] = cookie
    if referer:
        headers["Referer"] = referer
    last = None
    for i in range(retries):
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.getcode(), json.load(r)
        except urllib.error.HTTPError as e:
            try:
                return e.code, json.load(e)
            except Exception:
                return e.code, None
        except Exception as e:
            last = e
            time.sleep(0.6 * (i + 1))
    raise last


# ---------------------------------------------------------------- 목록 열거(비로그인 가능)
def enumerate_articles(cafe, menu, cookie=None):
    all_a, page = {}, 1
    cafe_name = ""
    while True:
        url = (f"{APIS}/cafe2/ArticleListV2dot1.json?search.clubid={cafe}"
               f"&search.menuid={menu}&search.queryType=lastArticle"
               f"&search.page={page}&search.perPage=50")
        code, d = http_get(url, cookie, f"https://cafe.naver.com/f-e/cafes/{cafe}/menus/{menu}")
        if code != 200 or not d:
            raise SystemExit(f"목록 조회 실패 (HTTP {code}). URL/카페ID를 확인하세요.")
        res = d["message"]["result"]
        cafe_name = res.get("cafeName", cafe_name)
        for a in res.get("articleList", []):
            all_a[a["articleId"]] = a
        if not res.get("hasNext"):
            break
        page += 1
        if page > 400:
            print("경고: 400페이지 초과 — 안전 중단", file=sys.stderr)
            break
        time.sleep(0.15)
    return cafe_name, all_a


# ---------------------------------------------------------------- 본문/댓글 fetch(로그인 필요)
def fetch_article(cafe, menu, aid, cookie):
    ref = f"https://cafe.naver.com/f-e/cafes/{cafe}/articles/{aid}?menuid={menu}"
    url = (f"{APIS}/cafe-articleapi/v3/cafes/{cafe}/articles/{aid}"
           f"?query=&menuId={menu}&boardType=L&useCafeId=true&requestFrom=A")
    code, d = http_get(url, cookie, ref)
    return code, d


def fetch_comments(cafe, aid, cookie):
    """v2 댓글 엔드포인트를 hasNext 따라 전량 수집(orderBy=asc)."""
    ref = f"https://cafe.naver.com/f-e/cafes/{cafe}/articles/{aid}"
    items, page = [], 1
    while True:
        url = (f"{APIS}/cafe-articleapi/v2/cafes/{cafe}/articles/{aid}"
               f"/comments/pages/{page}?requestFrom=A&orderBy=asc")
        code, d = http_get(url, cookie, ref)
        if code != 200 or not d:
            break
        r = d.get("result", d)
        block = r.get("comments") or {}
        items.extend(block.get("items", []) if isinstance(block, dict) else [])
        if not r.get("hasNext"):
            break
        page += 1
        if page > 40:
            break
        time.sleep(0.12)
    return items


# ---------------------------------------------------------------- HTML → 텍스트(원문 보존)
def html_to_text(html):
    if not html:
        return ""
    soup = BeautifulSoup(html, "lxml")
    for t in soup(["script", "style"]):
        t.decompose()
    for img in soup.find_all("img"):
        src = img.get("data-src") or img.get("src") or ""
        alt = img.get("alt") or ""
        img.replace_with(f"\n[이미지{(' '+alt) if alt else ''}: {src}]\n")
    for br in soup.find_all("br"):
        br.replace_with("\n")
    for td in soup.find_all(["td", "th"]):
        td.append(" | ")
    for tr in soup.find_all("tr"):
        tr.append("\n")
    for a in soup.find_all("a", href=True):
        href = a.get("href", "").strip()
        if href.startswith("http") and href not in a.get_text(strip=True):
            a.append(f" ({href})")
    for tag in soup.find_all(["p", "div", "li", "h1", "h2", "h3", "h4", "h5", "h6", "blockquote"]):
        tag.append("\n")
    container = soup.select_one(".se-main-container") or soup.body or soup
    text = container.get_text()
    out, blank = [], 0
    for ln in (l.rstrip() for l in text.split("\n")):
        if not ln.strip():
            blank += 1
            if blank <= 1:
                out.append("")
        else:
            blank = 0
            out.append(ln.strip())
    return "\n".join(out).strip()


def _txt(h):
    return BeautifulSoup(h, "lxml").get_text(" ", strip=True) if h else ""


def render_scrap(scrap):
    """레거시 [공유]/스크랩 글: article.scrap 에 본문이 있음."""
    if not scrap:
        return ""
    ch = scrap.get("contentHtml", "") or ""
    ce = scrap.get("contentElements") or []

    def render_el(e):
        if not isinstance(e, dict):
            return ""
        t, j = e.get("type"), (e.get("json") or {})
        if t == "IMAGE":
            return f"[이미지: {(j.get('image') or {}).get('url','')}]"
        if t == "LINK":
            title = htmlmod.unescape(j.get("truncatedTitle") or j.get("title") or "")
            desc = htmlmod.unescape(j.get("desc") or j.get("truncatedDesc") or "")
            link = j.get("linkUrl") or j.get("link") or ""
            dom = j.get("domain") or ""
            parts = [p for p in [title, desc, link, (f"({dom})" if dom else "")] if p]
            return "[링크: " + " — ".join(parts) + "]"
        return f"[{t}]"

    def repl(m):
        i = int(m.group(1))
        return render_el(ce[i]) if i < len(ce) else ""

    body = html_to_text(re.sub(r"\[\[\[CONTENT-ELEMENT-(\d+)\]\]\]", repl, ch))
    head = " ".join(x for x in [_txt(scrap.get("titleHtml")), _txt(scrap.get("linkHtml"))] if x)
    return (("[출처] " + head + "\n\n") if head else "") + body


def ts(ms):
    try:
        return datetime.fromtimestamp(int(ms) / 1000, KST).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""


ATTACH_FLAGS = [("attachImage", "이미지"), ("attachFile", "파일"), ("attachMovie", "동영상"),
                ("attachMusic", "음악"), ("attachPoll", "투표"), ("attachMap", "지도"),
                ("attachLink", "링크"), ("attachCalendar", "일정")]


def build_record(cafe, menu, aid, meta, content_json, comment_items):
    res = (content_json or {}).get("result", {})
    art = res.get("article")
    if not art:
        reason = (res.get("errorCode") or "") + " " + (res.get("reason") or "")
        return None, f"본문 없음 ({reason.strip() or 'unknown'})"

    writer = (art.get("writer") or {}).get("nick", "") or meta.get("writerNickname", "")
    body = html_to_text(art.get("contentHtml", "")) or render_scrap(art.get("scrap"))
    atts = res.get("attaches") or []
    img_urls = [a.get("url") for a in atts if a.get("type") == "I" and a.get("url")]
    img_urls = [u for u in img_urls if u not in body]
    if img_urls:
        body = (body + "\n\n" if body.strip() else "") + "[첨부 이미지]\n" + "\n".join(img_urls)

    att_label = ",".join(lab for k, lab in ATTACH_FLAGS if meta.get(k))
    fnames = [a.get("name") or a.get("fileName") for a in atts if a.get("name") or a.get("fileName")]
    if fnames:
        att_label = (att_label + " | " if att_label else "") + "; ".join(fnames)

    comments = []
    for c in (comment_items or []):
        kind = ("삭제됨" if c.get("isDeleted") else "답글" if c.get("isRef")
                else "작성자" if c.get("isArticleWriter") else "댓글")
        comments.append({
            "writer": (c.get("writer") or {}).get("nick", ""),
            "content": c.get("content", ""),
            "date": ts(c.get("updateDate")),
            "kind": kind,
        })

    rec = {
        "번호": aid,
        "제목": art.get("subject", ""),
        "작성자": writer,
        "작성일": ts(art.get("writeDate") or meta.get("writeDateTimestamp")),
        "조회수": art.get("readCount", meta.get("readCount", 0)),
        "댓글수": art.get("commentCount", meta.get("commentCount", 0)),
        "추천수": meta.get("likeItCount", 0),
        "공지여부": "공지" if art.get("isNotice") else "",
        "첨부": att_label,
        "본문": body,
        "댓글": "\n".join(f"[{c['kind']}] {c['writer']} ({c['date']}): {c['content']}" for c in comments),
        "원문링크": f"https://cafe.naver.com/f-e/cafes/{cafe}/articles/{aid}?menuid={menu}",
        "_comments": comments,
    }
    return rec, None


# ---------------------------------------------------------------- 엑셀
def write_excel(out, cafe, menu, cafe_name, articles):
    HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
    HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
    HEAD_AL = Alignment(horizontal="center", vertical="center", wrap_text=False)
    CELL = Alignment(vertical="top", wrap_text=True)
    NUM = Alignment(vertical="top", horizontal="center", wrap_text=False)
    thin = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    MAX = 32000

    def clip(v):
        return v[:MAX] + "\n...(생략)" if isinstance(v, str) and len(v) > MAX else v

    def sheet(ws, rows, cols, widths, nums, link=None):
        for c, name in enumerate(cols, 1):
            cell = ws.cell(1, c, name)
            cell.fill, cell.font, cell.alignment, cell.border = HEAD_FILL, HEAD_FONT, HEAD_AL, BORDER
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        for r, row in enumerate(rows, 2):
            for c, name in enumerate(cols, 1):
                cell = ws.cell(r, c, clip(row.get(name, "")))
                cell.border = BORDER
                cell.alignment = NUM if name in nums else CELL
                if link and name == link and row.get(name):
                    cell.hyperlink = row[name]
                    cell.font = Font(color="0563C1", underline="single", size=10)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
        ws.sheet_view.showGridLines = False

    comments = []
    for a in articles:
        for c in a.get("_comments", []):
            comments.append({"게시글번호": a["번호"], "게시글제목": a["제목"], "댓글작성자": c["writer"],
                             "댓글내용": c["content"], "댓글작성일": c["date"], "구분": c["kind"]})

    wb = Workbook()
    info = wb.active
    info.title = "정보"
    info.column_dimensions["A"].width = 14
    info.column_dimensions["B"].width = 95
    rows = [("카페명", cafe_name), ("카페ID", cafe), ("메뉴ID", menu),
            ("원본 URL", f"https://cafe.naver.com/f-e/cafes/{cafe}/menus/{menu}"),
            ("게시글 수", len(articles)), ("댓글 수", len(comments)),
            ("수집일", datetime.now(KST).strftime("%Y-%m-%d %H:%M")),
            ("비고", "제목/작성자/작성일/조회수/댓글수/추천수/본문/댓글/첨부/원문링크 전수 수집. "
                     "공유(스크랩)글은 출처링크·이미지 URL 포함.")]
    for r, (k, v) in enumerate(rows, 1):
        a = info.cell(r, 1, k); a.font = Font(bold=True); a.alignment = Alignment(vertical="top")
        info.cell(r, 2, v).alignment = Alignment(vertical="top", wrap_text=True)
    info.sheet_view.showGridLines = False

    sheet(wb.create_sheet("게시글"), articles,
          ["번호", "제목", "작성자", "작성일", "조회수", "댓글수", "추천수", "공지여부", "첨부", "본문", "댓글", "원문링크"],
          [8, 42, 18, 18, 8, 8, 8, 8, 24, 85, 60, 42],
          {"번호", "조회수", "댓글수", "추천수", "공지여부"}, link="원문링크")
    sheet(wb.create_sheet("댓글"), comments,
          ["게시글번호", "게시글제목", "댓글작성자", "댓글내용", "댓글작성일", "구분"],
          [12, 42, 20, 70, 18, 10], {"게시글번호", "구분"})
    wb.save(out)
    return len(comments)


# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--url", required=True)
    ap.add_argument("--cookie", required=True, help="쿠키 헤더 파일 경로")
    ap.add_argument("--out", default="")
    ap.add_argument("--workers", type=int, default=8)
    args = ap.parse_args()

    cafe, menu = parse_url(args.url)
    cookie = open(args.cookie).read().strip()
    if "NID_SES" not in cookie or "NID_AUT" not in cookie:
        print("경고: 쿠키에 NID_AUT/NID_SES 가 없습니다 — 본문이 안 보일 수 있습니다.", file=sys.stderr)

    print(f"[1/4] 목록 열거… (카페 {cafe} / 메뉴 {menu})")
    cafe_name, meta_by_id = enumerate_articles(cafe, menu, cookie)
    ids = sorted(meta_by_id)
    print(f"      카페: {cafe_name} | 게시글 {len(ids)}건")

    print(f"[2/4] 본문+댓글 병렬 수집… (worker {args.workers})")
    results, errors = {}, []

    def work(aid):
        code, cj = fetch_article(cafe, menu, aid, cookie)
        citems = fetch_comments(cafe, aid, cookie)
        rec, err = build_record(cafe, menu, aid, meta_by_id.get(aid, {}), cj, citems)
        return aid, code, rec, err

    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = {ex.submit(work, a): a for a in ids}
        done = 0
        for f in as_completed(futs):
            aid, code, rec, err = f.result()
            done += 1
            if rec:
                results[aid] = rec
            else:
                errors.append((aid, code, err))
            if done % 10 == 0 or done == len(ids):
                print(f"      {done}/{len(ids)}")

    articles = [results[a] for a in sorted(results, key=lambda x: results[x]["작성일"], reverse=True)]

    print("[3/4] 정합성 점검…")
    empty = [a["번호"] for a in articles if not a["본문"].strip()]
    print(f"      수집 {len(articles)}건 / 실패 {len(errors)}건 / 본문없음 {len(empty)}건")
    if errors:
        for aid, code, err in errors[:20]:
            print(f"      - {aid}: HTTP {code} {err}")
        if any(code in (401, 403) for _, code, _ in errors):
            print("      ※ 401/403 다수 → 이 계정이 카페 '회원'이 아니거나 쿠키가 만료됐을 수 있습니다.", file=sys.stderr)

    out = args.out or f"{cafe_name or cafe}_{menu}_크롤링_{datetime.now(KST).strftime('%Y%m%d')}.xlsx"
    out = os.path.abspath(out)
    ncom = write_excel(out, cafe, menu, cafe_name, articles)
    print(f"[4/4] 엑셀 저장 완료 → {out}")
    print(f"      게시글 {len(articles)} · 댓글 {ncom}")
    # 실패가 전부(수집 0) 면 실패 코드로 종료
    if not articles:
        sys.exit(3)


if __name__ == "__main__":
    main()
